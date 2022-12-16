# Fridge Project by: Gary Soo 
# Sources: Mr. Cozort 
# Datetime: https://www.geeksforgeeks.org/get-current-date-using-python/#:~:text=now(),defined%20under%20the%20DateTime%20module.
# Openpyxl: https://www.geeksforgeeks.org/python-writing-excel-file-using-openpyxl-module/
# Tkinter with Openpyxl: https://www.youtube.com/watch?v=l6-HG0FJPsQ
# Openpyxl: https://openpyxl.readthedocs.io/en/stable/
# Sort Function: https://www.educative.io/answers/how-to-sort-a-list-of-tuples-in-python-using-lambda
# Sort Function: https://www.freecodecamp.org/news/python-sort-list-how-to-order-by-descending-or-ascending/
# Pandas: https://pandas.pydata.org/docs/reference/api/pandas.DataFrame.sort_values.html
# Pandas: https://www.geeksforgeeks.org/add-column-names-to-dataframe-in-pandas/
# Pip: https://pypi.org/

# Fridge Project 2022
# 1 Input Item, Category, Expiration Date
# 2 Add Data to a Dictionary
# 3 Import Data to txt file
# 4 Sort Dictionary by Expiration Date
# 5 Visual Component

# using os to create file, using tkinter for window
# will use sys to exit program
# using openpyxl to read and write in excel
# using datetime to grab current date
# using uuid to generate random uuid
# using pathlib to locate file
from tkinter import *
from tkinter import ttk
from tkinter import messagebox
import uuid
import os
import sys 
from pathlib import Path
import pprint
import datetime as dt
import openpyxl
import pandas as pd
import pip

#Auto install libraries
pip.main(["install", "pandas"])
pip.main(["install", "tkinter"])
pip.main(["install", "openpyxl"])
pip.main(["install", "tkinter"])
pip.main(["install", "pathlib"])
pip.main(["install", "datetime"])

print(str(Path.cwd()))
os.chdir(Path.cwd())

# Check to see if file exist, if not create new file
if os.path.exists('FridgeData.xlsx'):
    print('File exists')
else:
    print('File does not exist')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    sheet = wb["Sheet1"]
    wb.save('FridgeData.xlsx')
    print('File created')

# Check to see if sheet exist, if not create new sheet
wb = openpyxl.load_workbook('FridgeData.xlsx')

if 'Sheet1' in wb.sheetnames:
    print('Sheet1 exists')
else:
    print('Sheet1 does not exist')
    wb.create_sheet('Sheet1')
    print('Sheet1 created')

wb.save('FridgeData.xlsx')

#Create an instance of tkinter frame
win = Tk()
#Set the geometry of tkinter frame
win.geometry("750x400")
win.title("Sort Food by Expiration Date")

#Creates a Frame
frame = LabelFrame(win, width= 400, height= 180, bd=5)
frame.pack()
#Stop the frame from propagating the widget to be shrink or fit
frame.pack_propagate(False)

#Create an Entry widget in the Frame
fooditem = ttk.Entry(frame, width= 40)
fooditem.insert(INSERT, "Enter Food")
fooditem.pack()

category = ttk.Entry(frame, width= 40)
category.insert(INSERT, "Enter Category")
category.pack()

ExDate = ttk.Entry(frame, width= 40)
ExDate.insert(INSERT, "Enter Expiry Date as mm/dd/yy")
ExDate.pack()

# Adds data to excel file
def add_item():
   # calendar convert date into integer for sorting
   date = str(ExDate.get())
   day = date[3:5]
   print("entered day:"+ day)
   month = date[0:2]
   print("entered month:" + month)
   year = date[6:8]
   print("entered year:" + year)

# Check if entry is integer or not
#    if isinstance(day, int) == False:
#     messagebox.showinfo("Try Again", "Date not in correct format")

   numdate = int(day) + int(month) * 30 + int(year) * 365
   print("this is your date: " + str(numdate))

   # turn today's date into an integer
   now = dt.date.today()
   todayyear = now.strftime("%Y")
   print("year:", todayyear)
   todaymonth = now.strftime("%m")
   print("month:", todaymonth)
   todayday = now.strftime("%d")
   print("day:", todayday)
   numtoday = int(todayday[0:2]) + int(todaymonth[0:2]) * 30 + int(todayyear[2:4]) * 365
   print("test" + todayyear[2:4])
   print("this is today's date: " + str(numtoday))
   # days until ExDate
   daystill = numdate - numtoday
   print("days until expiration " + str(daystill))
   label = Label(frame, text = "Command Recieved" , font= ('Times New Roman', 9, 'italic'))

   # Get fooditems, category, and ExDate from entry boxes
   fooditem_data = fooditem.get()
   category_data = category.get()
   ExDate_data = daystill
      
   # Open Excel File
   wb = openpyxl.load_workbook("FridgeData.xlsx")
   sheet = wb["Sheet1"]

   # Add data without erasing other rows
   empty_row = sheet.max_row + 1
   sheet.cell(row = empty_row, column = 1).value = fooditem_data
   sheet.cell(row = empty_row, column = 2).value = category_data
   sheet.cell(row = empty_row, column = 3).value = ExDate_data
   wb.save("FridgeData.xlsx")

   # Clear entry boxes
   fooditem.delete(0, 'end')
   category.delete(0, 'end')
   ExDate.delete(0, 'end')
   label.pack(pady=30)
   messagebox.showinfo("Success", "fooditems, category, and ExDate added to excel file")

# Sorts Data in Excel File
def sort():
   # Open excel file
   wb = openpyxl.load_workbook("FridgeData.xlsx")
   sheet = wb["Sheet1"]

   # Get excel file, open sheet 1, then sort the 3nd column from smallest to largest
   xl = pd.ExcelFile("FridgeData.xlsx")
   df = xl.parse("Sheet1")
   df.columns =['Food', 'Category', 'Days until Expiration Date']
   final_result = df.sort_values(df.columns[2], ascending=True)
   # df = df.sort_values("C", ascending=False)

   # Creates Excel File with sorted data
   print()
   print()
   print()
   print(final_result)
   final_result.to_excel('SortedFridgeData.xlsx')
   messagebox.showinfo("Success", "Data Sort Complete")

# Clears excel file
def clear_data():
    # Open excel file
    wb = openpyxl.load_workbook("FridgeData.xlsx")
    sheet = wb["Sheet1"]

    # Loop through rows in column A, B, and C
    for r in range(1, sheet.max_row + 1):
        # Clear fooditems, category, and ExDate
        sheet.cell(row = r, column = 1).value = None
        sheet.cell(row = r, column = 2).value = None
        sheet.cell(row = r, column = 3).value = None

    # Save the file
    wb.save("FridgeData.xlsx")

    # Display a message box when excel file is cleared
    messagebox.showinfo("Success", "Excel file cleared")

#Create a Button
ttk.Button(win, text= "Add Item", command= add_item).pack(pady=20)
ttk.Button(win, text= "Sort", command = sort).pack(pady=20)
ttk.Button(win, text= "Delete Data", command= clear_data).pack(pady=20)
win.mainloop()